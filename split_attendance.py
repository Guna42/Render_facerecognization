import os
from openpyxl import load_workbook, Workbook

def split_attendance(current_session_id, known_faces_db):
    # File paths
    attendance_file = r"C:\Users\GUNA\Videos\Recordings\Attendence_Automation\attendance.xlsx"
    present_file = r"C:\Users\GUNA\Videos\Recordings\Attendence_Automation\present_students.xlsx"
    absent_file = r"C:\Users\GUNA\Videos\Recordings\Attendence_Automation\absent_students.xlsx"

    # Get present students from attendance.xlsx for current session
    present_students = set()
    if os.path.exists(attendance_file):
        wb = load_workbook(attendance_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == current_session_id:
                present_students.add(row[1])  # Name is in column 2
        wb.close()

    # Get all students from known_faces_db
    all_students = {details["name"]: details["email"] for details in known_faces_db.values()}

    # Absentees = all students - present students
    absent_students = {name: email for name, email in all_students.items() if name not in present_students}

    # Write present students
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Email", "Status"])
    for name in present_students:
        email = all_students[name]
        ws.append([name, email, "Present"])
    wb.save(present_file)

    # Write absent students
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Email", "Status"])
    for name, email in absent_students.items():
        ws.append([name, email, "Absent"])
    wb.save(absent_file)

    print(f"Attendance split: {len(present_students)} present, {len(absent_students)} absent")

if __name__ == "__main__":
    # This is just for testing standalone; we'll call it from face_recognition_live.py
    sample_session_id = "test_session"
    sample_db = {
        "kartheek.jpg": {"name": "Kartheek Sanka", "email": "kartheek.reddy@example.com"},
        "sumanth.jpg": {"name": "Sumanth ", "email": "sumanth.kumar@example.com"}
    }
    split_attendance(sample_session_id, sample_db)