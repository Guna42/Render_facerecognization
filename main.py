import smtplib
import datetime
import time
import os
from openpyxl import load_workbook

# File path
ABSENT_FILE = r"C:\Users\GUNA\Videos\Recordings\Attendence_Automation\absent_students.xlsx"

# Email credentials
MY_EMAIL = "srmap999@gmail.com"
PASSWORD = "fblq qwax wtci aijl"

def send_absentee_emails():
    print("Starting send_absentee_emails function")
    
    # Check if file exists
    if not os.path.exists(ABSENT_FILE):
        print(f"Error: {ABSENT_FILE} not found")
        raise FileNotFoundError(f"{ABSENT_FILE} not found")

    # Read Excel file
    try:
        wb = load_workbook(ABSENT_FILE)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        print(f"Loaded {ABSENT_FILE} with {len(rows)} rows")
        if not rows:
            print("No absentees found in Excel")
            wb.close()
            return
    except Exception as e:
        print(f"Error reading {ABSENT_FILE}: {e}")
        wb.close()
        raise

    # Get date and time
    today = datetime.datetime.now()
    today_date = today.strftime("%Y-%m-%d")
    current_time = today.strftime("%H:%M:%S")

    for row in rows:
        name, email, status = row
        if status.strip().lower() == "absent":
            # Read letter template
            try:
                with open(r"C:\Users\GUNA\Videos\Recordings\Attendence_Automation\templates\letter_1.txt") as file:
                    content = file.read()
                    message = content.replace("NAME", name).replace("DATE", today_date).replace("TIME", current_time)
                print(f"Prepared email for {name} ({email})")
            except Exception as e:
                print(f"Error reading letter template: {e}")
                raise

            # Send email
            try:
                with smtplib.SMTP("smtp.gmail.com", 587) as connect:
                    connect.starttls()
                    connect.login(MY_EMAIL, PASSWORD)
                    connect.sendmail(
                        from_addr=MY_EMAIL,
                        to_addrs=email,
                        msg=f"Subject: IMPORTANT Attendance Notification\n\n{message}"
                    )
                print(f"Email successfully sent to {name} ({email})")
                time.sleep(2)  # Avoid rate limits
            except Exception as e:
                print(f"Failed to send email to {name} ({email}): {e}")
                raise

    wb.close()
    print("All emails processed!")

if __name__ == "__main__":
    send_absentee_emails()